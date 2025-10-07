import io
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from thefuzz import fuzz

import dictionary_matcher
import value_dictionary_handler
from utils import normalize_header, get_col_from_cell
from local_geocoding_service import apply_post_processing

def _apply_manual_rules(source_ws, template_ws, rules, s_start_row, t_start_row, used_source_cols, used_template_cols,
                        visible_rows_only):
    s_end_row = source_ws.max_row
    for rule in rules:
        s_col_letter = rule.get('s_col') or get_col_from_cell(rule.get('source_cell'))
        t_col_letter = rule.get('t_col') or rule.get('template_col')

        if not s_col_letter or not t_col_letter:
            continue

        s_col_idx, t_col_idx = column_index_from_string(s_col_letter), column_index_from_string(t_col_letter)

        if s_col_idx in used_source_cols or t_col_idx in used_template_cols:
            continue

        target_row_counter = 0
        for r_idx in range(s_start_row + 1, s_end_row + 1):
            if visible_rows_only and source_ws.row_dimensions[r_idx].hidden:
                continue

            source_cell = source_ws.cell(row=r_idx, column=s_col_idx)
            target_cell = template_ws.cell(row=t_start_row + 1 + target_row_counter, column=t_col_idx)

            target_cell.value = source_cell.value
            if source_cell.hyperlink:
                target_cell.hyperlink = source_cell.hyperlink.target
                target_cell.style = "Hyperlink"
            target_row_counter += 1

        used_source_cols.add(s_col_idx)
        used_template_cols.add(t_col_idx)


def _apply_dictionary_matching(source_ws, template_ws, s_start_row, t_start_row, used_source_cols, used_template_cols,
                               visible_rows_only):
    reverse_dictionary = dictionary_matcher.get_reverse_dictionary()
    s_headers = {c.column: normalize_header(c.value) for c in source_ws[s_start_row] if c.value}
    t_headers = {c.column: normalize_header(c.value) for c in template_ws[t_start_row] if c.value}
    s_end_row = source_ws.max_row

    for s_col_idx, s_norm_h in s_headers.items():
        if s_col_idx in used_source_cols: continue
        canonical_name = reverse_dictionary.get(s_norm_h)
        if not canonical_name: continue
        normalized_canonical = normalize_header(canonical_name)

        for t_col_idx, t_norm_h in t_headers.items():
            if t_col_idx in used_template_cols: continue
            if t_norm_h == normalized_canonical:
                target_row_counter = 0
                for r_idx in range(s_start_row + 1, s_end_row + 1):
                    if visible_rows_only and source_ws.row_dimensions[r_idx].hidden:
                        continue

                    source_cell = source_ws.cell(row=r_idx, column=s_col_idx)
                    target_cell = template_ws.cell(row=t_start_row + 1 + target_row_counter, column=t_col_idx)

                    target_cell.value = source_cell.value
                    if source_cell.hyperlink:
                        target_cell.hyperlink = source_cell.hyperlink.target
                        target_cell.style = "Hyperlink"
                    target_row_counter += 1

                used_source_cols.add(s_col_idx)
                used_template_cols.add(t_col_idx)
                break


def _apply_auto_matching(source_ws, template_ws, s_start_row, t_start_row, used_source_cols, used_template_cols,
                         task_id, task_statuses, visible_rows_only):
    s_headers = {c.column: normalize_header(c.value) for c in source_ws[s_start_row] if c.value}
    t_headers = {c.column: normalize_header(c.value) for c in template_ws[t_start_row] if c.value}
    s_end_row = source_ws.max_row

    auto_source_headers = {k: v for k, v in s_headers.items() if k not in used_source_cols}
    auto_template_headers = {k: v for k, v in t_headers.items() if k not in used_template_cols}
    auto_matches = {}

    for s_col, s_norm in auto_source_headers.items():
        best_match, best_score = None, 0
        for t_col, t_norm in auto_template_headers.items():
            score = fuzz.ratio(s_norm, t_norm)
            if score > best_score:
                best_score, best_match = score, t_col
        if best_score > 75:
            auto_matches[s_col] = best_match
            auto_template_headers = {k: v for k, v in auto_template_headers.items() if k != best_match}

    rows_to_process = []
    for r_idx in range(s_start_row + 1, s_end_row + 1):
        if not (visible_rows_only and source_ws.row_dimensions[r_idx].hidden):
            rows_to_process.append(source_ws[r_idx])

    for i, row in enumerate(rows_to_process):
        for s_col, t_col in auto_matches.items():
            source_cell = row[s_col - 1]
            target_cell = template_ws.cell(row=t_start_row + 1 + i, column=t_col)

            target_cell.value = source_cell.value
            if source_cell.hyperlink:
                target_cell.hyperlink = source_cell.hyperlink.target
                target_cell.style = "Hyperlink"

        if (i + 1) % 100 == 0:
            task_statuses[task_id]['status'] = f'Автоматическое копирование: строка {i + 1}'


def _apply_value_dictionary(worksheet, task_id, task_statuses):
    reverse_rules_map = value_dictionary_handler.get_reverse_lookup_map()
    if not reverse_rules_map:
        return

    task_statuses[task_id]['status'] = 'Выполняю замену по словарю значений...'
    replacements_count = 0

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                replacement = reverse_rules_map.get(cell.value)
                if replacement is not None:
                    cell.value = replacement
                    replacements_count += 1

    task_statuses[task_id]['status'] = f'Замена по словарю завершена ({replacements_count} замен)'


def process_excel_hybrid(task_id, source_file_obj, template_file_obj, ranges, template_rules, private_rules,
                         post_function, original_template_filename, task_statuses, visible_rows_only):
    """
    Основная функция обработки файлов Excel.
    """
    try:
        task_statuses[task_id] = {
            'progress': 5,
            'status': 'Подготовка...',
            'template_filename': original_template_filename
        }
        source_wb = load_workbook(filename=source_file_obj)
        source_ws = source_wb.active

        is_macro_enabled = original_template_filename.lower().endswith('.xlsm')
        template_wb = load_workbook(filename=template_file_obj, keep_vba=is_macro_enabled)
        template_ws = template_wb.active

        s_start_row, t_start_row = ranges['s_start_row'], ranges['t_start_row']
        used_source_cols, used_template_cols = set(), set()

        task_statuses[task_id]['status'] = 'Выполняю частные правила...'
        _apply_manual_rules(source_ws, template_ws, private_rules, s_start_row, t_start_row, used_source_cols,
                            used_template_cols, visible_rows_only)

        task_statuses[task_id]['status'] = 'Применяю правила из шаблона...'
        _apply_manual_rules(source_ws, template_ws, template_rules, s_start_row, t_start_row, used_source_cols,
                            used_template_cols, visible_rows_only)

        task_statuses[task_id]['status'] = 'Проверяю по словарю...'
        _apply_dictionary_matching(source_ws, template_ws, s_start_row, t_start_row, used_source_cols,
                                   used_template_cols, visible_rows_only)

        task_statuses[task_id]['status'] = 'Ищу автоматические совпадения...'
        _apply_auto_matching(source_ws, template_ws, s_start_row, t_start_row, used_source_cols, used_template_cols,
                             task_id, task_statuses, visible_rows_only)

        _apply_value_dictionary(template_ws, task_id, task_statuses)

        task_statuses[task_id]['status'] = 'Запускаю пост-обработку...'
        apply_post_processing(task_id, template_wb, t_start_row, post_function, task_statuses)

        task_statuses[task_id]['status'] = 'Сохраняю результат...'
        processed_file_obj = io.BytesIO()
        template_wb.save(processed_file_obj)
        processed_file_obj.seek(0)
        source_wb.close()
        template_wb.close()
        task_statuses[task_id].update({'progress': 100, 'status': 'Готово!', 'result_file': processed_file_obj})

    except Exception as e:
        task_statuses[task_id].update({'progress': 100, 'status': f"Ошибка: {e}", 'result_file': None})